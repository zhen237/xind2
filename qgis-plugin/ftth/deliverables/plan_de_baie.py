# -*- coding: utf-8 -*-
"""
Plan_de_Baie (机柜熔接盘图) 生成器 (plan_de_baie.py)
====================================================

按主办方官方样本 (EJA01_MRJ01_Plan_de_Baie.xlsx, sheet "Plan Baie") 的逐格版式生成
机柜熔接盘占用表：

  - DENOMINATION / LOCALISATION 表头
  - "Baie {PM}" 标题横幅 (合并 B6:AB7)
  - 24 芯位列 (D..AA) + Tubes(AC) + REFERENCE CABLE(AD) 表头
  - STK 储备缆行 (stockage)
  - TDI 抽屉行 (每个抽屉 = 1 条缆，子行 = 该缆的束管对 Tube x-Tube y，D..AA 填满 1..24 芯位)

重要(v1 同源限制):
  本表核心是「熔接盘/光纤占用矩阵」，其中**每芯位占用、束管→缆段扇出**属于营运商熔接计划，
  8 个 Shape 图层并不携带。故:
    * 真实可推导值(全部填真值): 机柜 PM 编码、局站地名、缆段 CODE、缆段总数、箱体归属
    * 确定性占位(代码内明确标注 PLACEHOLDER): 芯位逐格占用、STK 的 Tube 12 固定标注、
      抽屉子行束管对划分。这些占位遵循官方版式结构，待真实熔接计划数据补入即可替换。
"""

from __future__ import annotations

MAX_DRAWERS = 12          # 单柜最多绘制的抽屉数(超过则截断展示)
MAX_STK = 5               # 储备缆行数
TUBES_PER_CABLE = 12      # FTTH 标准 12 束管光缆


def _mrj_suffix(project, pm_code: str) -> str:
    pms = project.pm_codes()
    idx = pms.index(pm_code) + 1 if pm_code in pms else 1
    return f"MRJ{idx:02d}"


def build_plan_de_baie(project, pm_code: str) -> dict:
    """构造 Plan_de_Baie 的结构化数据。"""
    mrj = _mrj_suffix(project, pm_code)
    localisation = project.dominant_localisation() or "NA"

    cables = project.connected_cables_for_pm(pm_code)
    # 抽屉: 每条缆一个抽屉(上限 MAX_DRAWERS)
    drawer_cables = cables[:MAX_DRAWERS]
    # 储备: 其余缆取前 MAX_STK 条作 stockage
    stk_cables = cables[MAX_DRAWERS:MAX_DRAWERS + MAX_STK]

    tdi_rows = []
    for di, cable in enumerate(drawer_cables, start=1):
        drawer_id = f"TDI{di:02d}-{mrj}"
        n_sub = max(1, (TUBES_PER_CABLE + 1) // 2)  # 12 管 -> 6 子行
        for si in range(n_sub):
            tube_lo = si * 2 + 1
            tube_hi = min(tube_lo + 1, TUBES_PER_CABLE)
            tdi_rows.append({
                "drawer": drawer_id,
                "sub": chr(ord("A") + si),
                "tube_range": f"Tube {tube_lo}-Tube {tube_hi}",
                "cable": cable,
            })

    stk_rows = []
    for si, cable in enumerate(stk_cables, start=1):
        stk_rows.append({
            "label": f"{pm_code}-STK{si:02d}",
            "tube": "Tube 12",          # PLACEHOLDER: 储备缆统一标 Tube 12 (同官方样本)
            "cable": cable,
            "col_offset": 2 * si,       # PLACEHOLDER: Tube 12 标注所在芯位列偏移
        })

    return {
        "pm": pm_code,
        "mrj": mrj,
        "localisation": localisation,
        "title": f"Baie {pm_code}",
        "stk_rows": stk_rows,
        "tdi_rows": tdi_rows,
        "n_cables_total": len(cables),
    }


def export_plan_de_baie_xlsx(project, out_path: str, pm_code: str) -> str:
    """写出 Plan_de_Baie xlsx (官方 "Plan Baie" sheet 版式)。返回文件路径。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = build_plan_de_baie(project, pm_code)
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan Baie"

    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)

    # ---- 表头 ----
    ws["C2"] = "DENOMINATION:"
    ws["D2"] = data["pm"]
    ws["C3"] = "LOCALISATION:"
    ws["D3"] = data["localisation"]
    ws["C2"].font = bold
    ws["C3"].font = bold

    # ---- 标题横幅 ----
    ws["B6"] = data["title"]
    ws["B6"].font = Font(bold=True, size=14)
    ws.merge_cells("B6:AB7")
    ws["B6"].alignment = Alignment(horizontal="center", vertical="center")

    # ---- 24 芯位表头 (row 9 标签, row 10 序号) ----
    ws["C9"] = "Position\nPlateau"
    ws["C9"].font = bold
    ws["C9"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("C9:C10")
    ws["AC9"] = "Tubes"
    ws["AC9"].font = bold
    ws["AD9"] = "REFERENCE CABLE"
    ws["AD9"].font = bold
    for i, col in enumerate(range(4, 28), start=1):  # D..AA = 1..24
        c = ws.cell(row=10, column=col, value=i)
        c.font = bold
        c.alignment = Alignment(horizontal="center")
    ws.merge_cells("AC9:AC10")
    ws.merge_cells("AD9:AD10")

    # ---- STK 储备缆行 ----
    r = 11
    for stk in data["stk_rows"]:
        ws.cell(row=r, column=3, value=stk["label"]).font = bold
        ws.merge_cells(start_row=r, start_column=3, end_row=r + 1, end_column=3)
        col = 4 + (stk["col_offset"] % 24)
        ws.cell(row=r, column=col, value=stk["tube"])
        ws.cell(row=r, column=29, value=stk["tube"])  # AC
        ws.cell(row=r, column=30, value=stk["cable"])  # AD
        ws.merge_cells(start_row=r, start_column=30, end_row=r + 1, end_column=30)
        for cc in range(3, 31):
            ws.cell(row=r, column=cc).border = border
            ws.cell(row=r + 1, column=cc).border = border
        r += 2

    # ---- TDI 抽屉行 ----
    cur_drawer = None
    drawer_first_row = None
    for tdi in data["tdi_rows"]:
        if tdi["drawer"] != cur_drawer:
            # 关闭上一个抽屉的 AD 合并
            if cur_drawer is not None and drawer_first_row is not None:
                ws.merge_cells(start_row=drawer_first_row, start_column=30,
                               end_row=r - 1, end_column=30)
            cur_drawer = tdi["drawer"]
            drawer_first_row = r
        label = f"{tdi['drawer']}-{tdi['sub']}"
        ws.cell(row=r, column=3, value=label).font = bold
        ws.merge_cells(start_row=r, start_column=3, end_row=r + 1, end_column=3)
        for col in range(4, 28):
            ws.cell(row=r, column=col, value=col - 3)  # D=1 .. AA=24
        ws.cell(row=r, column=29, value=tdi["tube_range"])  # AC
        ws.cell(row=r, column=30, value=tdi["cable"])      # AD
        for cc in range(3, 31):
            ws.cell(row=r, column=cc).border = border
            ws.cell(row=r + 1, column=cc).border = border
        r += 2
    if cur_drawer is not None and drawer_first_row is not None:
        ws.merge_cells(start_row=drawer_first_row, start_column=30,
                       end_row=r - 1, end_column=30)

    # ---- 列宽 ----
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["AC"].width = 16
    ws.column_dimensions["AD"].width = 24
    for col in range(4, 28):
        ws.column_dimensions[get_column_letter(col)].width = 4

    wb.save(out_path)
    return out_path
