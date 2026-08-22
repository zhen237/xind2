# -*- coding: utf-8 -*-
"""
Synoptique (系统/光路连接图) 生成器 (synoptique.py)
====================================================

按主办方官方样本 (EJA01_MRJ01_Synoptique.xlsx, sheet "SYNO_MRJxx") 的逐格版式生成
各 PM 的光路系统图：

  - B5 标题 "{PM} - {MRJ}"
  - B6 "{PRISES} PRISES"  (该 PM 下属箱体住户总和)
  - B7.. Tiroir 抽屉列表 (MRJxxTDIyy | Tiroir 144FO)
  - B13 局站地名
  - 每个箱体一个块: 箱体 CODE + 住户数, 类型, 12 束管(Tube[Fx à Fy])→缆段映射, RESERVE/DEPLOYE/NA/CHA 标记

重要(v1 同源限制):
  本表核心是「束管→缆段扇出 + 熔接标记」的营运商熔接计划，8 个 Shape 图层并不携带逐束管/逐纤芯
  的扇出细节。故:
    * 真实可推导值(全部填真值): PM/MRJ、住户总和(PRISES)、抽屉列表、箱体 CODE、箱体类型、
      住户数、缆段 CODE(来自 route_for_boite 真实路由)、路由总长
    * 确定性占位(代码内明确标注 PLACEHOLDER): 12 束管→缆段的轮转分配、每束管纤维区间[F1 à F12]、
      T 列注解(类型/敷设方式)、RESERVE/DEPLOYE/NA/CHA 标记。结构遵循官方版式，待真实熔接计划补入即替换。
"""

from __future__ import annotations

import re


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """openpyxl sheet 名限制: <=31 字符，且不能含 []:*?/\\ 。做截断+清洗并去重。"""
    clean = re.sub(r"[\[\]:*?/\\]", "_", name)[:31]
    if not clean:
        clean = "Sheet"
    base = clean
    i = 1
    while clean in used:
        suffix = f"_{i}"
        clean = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(clean)
    return clean


TUBES = 12  # FTTH 标准 12 束管
FIBERS_PER_TUBE = 12
MAX_TIROIRS = 8


def _mrj_suffix(project, pm_code: str) -> str:
    pms = project.pm_codes()
    idx = pms.index(pm_code) + 1 if pm_code in pms else 1
    return f"MRJ{idx:02d}"


def build_synoptique(project, pm_code: str) -> dict:
    mrj = _mrj_suffix(project, pm_code)
    boites = project.boites_for_pm(pm_code)
    total_logements = sum(project.logements_of_boite(b) for b in boites)
    # PRISES 优先取 ZPM.NB_PRISES(与官方"516 PRISES"语义一致)，否则回退累加住户数
    zpm = project.zpm.get(pm_code, {})
    prises_raw = _s(zpm.get("NB_PRISES"))
    prises = total_logements
    if prises_raw:
        try:
            prises = int(float(prises_raw))
        except (TypeError, ValueError):
            pass
    cables = project.connected_cables_for_pm(pm_code)
    n_tiroirs = max(1, min(len(cables), MAX_TIROIRS))

    blocks = []
    for boite in boites:
        b = project.boites.get(boite, {})
        btype = (b.get("TYPE") or "").strip() or "BPE"
        logements = project.logements_of_boite(boite)
        route = project.route_for_boite(boite)
        boite_cables = [h[0] for h in (route or {}).get("hops", []) if h[0]]
        if not boite_cables:
            boite_cables = [b.get("CABLE_AMON")] if (b.get("CABLE_AMON") or "").strip() else []
        n = len(boite_cables)
        tubes = []
        for t in range(1, TUBES + 1):
            if n == 0:
                cable = "Passage"          # PLACEHOLDER: 无上游缆段
            else:
                cable = boite_cables[(t - 1) % n]
            annot = ""
            if cable not in ("Passage", "RESERVE"):
                cab = project.cables.get(cable, {})
                annot = _s(cab.get("MODE_POSE") or cab.get("TYPE_CABLE"))
            tubes.append({
                "tube": t,
                "fibers": FIBERS_PER_TUBE,
                "cable": cable,
                "annot": annot,           # PLACEHOLDER: T 列注解(敷设方式/类型)
            })
        blocks.append({
            "boite": boite,
            "type": btype,
            "logements": logements,
            "tubes": tubes,
            "reserve_brins": TUBES - n * 2 if n else TUBES,  # PLACEHOLDER
        })

    return {
        "pm": pm_code,
        "mrj": mrj,
        "prises": total_logements,
        "localisation": project.dominant_localisation() or "NA",
        "n_tiroirs": n_tiroirs,
        "blocks": blocks,
    }


def export_synoptique_xlsx(project, out_path: str = "", pm_code: str = "", workbook=None):
    """写出 Synoptique 系统图。

    如果传入 workbook，则把数据追加为一张新 sheet 而不保存；
    否则创建新工作簿并保存到 out_path。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    data = build_synoptique(project, pm_code)
    if workbook is None:
        wb = Workbook()
        save_path = out_path
    else:
        wb = workbook
        save_path = None

    used = set(wb.sheetnames)
    sheet_title = _safe_sheet_name(f"Syno_{data['mrj']}_{pm_code}", used)
    ws = wb.create_sheet(title=sheet_title)

    bold = Font(bold=True)
    ws["B5"] = f"{data['pm']} - {data['mrj']}"
    ws["B5"].font = Font(bold=True, size=13)
    ws["B6"] = f"{data['prises']} PRISES"
    ws["B6"].font = bold
    for k in range(1, data["n_tiroirs"] + 1):
        ws.cell(row=6 + k, column=2, value=f"{data['mrj']}TDI{k:02d}")
        ws.cell(row=6 + k, column=3, value="Tiroir 144FO")
    ws.cell(row=13, column=2, value=data["localisation"])

    r = 14
    for blk in data["blocks"]:
        # 块首行: 箱体 + 住户数
        ws.cell(row=r, column=26, value=blk["boite"]).font = bold   # Z
        ws.cell(row=r, column=27, value=f"{blk['logements']} LOGEMENTS")  # AA
        r += 1
        ws.cell(row=r, column=26, value=blk["type"])  # Z 类型
        r += 1
        for td in blk["tubes"]:
            ws.cell(row=r, column=20, value=td["annot"])  # T 注解(PLACEHOLDER)
            ws.cell(row=r, column=22, value="Tiroir 1")   # V
            ws.cell(row=r, column=26,
                    value=f"Tube {td['tube']}[F1 à F{td['fibers']}]")  # Z
            ws.cell(row=r, column=27, value=td["cable"])  # AA
            r += 1
        ws.cell(row=r, column=26, value=f"{max(0, blk['reserve_brins'])} Brins RESERVE")
        r += 1
        ws.cell(row=r, column=26, value="DEPLOYE")
        r += 1
        ws.cell(row=r, column=26, value="NA")
        r += 1
        ws.cell(row=r, column=26, value=f"CHA-{blk['boite']}")
        r += 2  # 块间空一行

    if save_path:
        wb.save(save_path)
        return save_path
    return sheet_title
