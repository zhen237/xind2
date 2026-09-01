# -*- coding: utf-8 -*-
"""
FTTH 交付物导出运行器 (export_runner.py)
=========================================

封装「装载 -> 生成两表 -> 写 xlsx」全流程，供:
  - 离线校验脚本调用 (export_from_dbf)
  - QGIS 插件按钮调用 (export_from_qgis)
"""

from __future__ import annotations

import os

import openpyxl

from .loader import load_dbf, load_qgis
from .deliverables import (
    export_plans_de_boite_xlsx,
    export_routes_optiques_xlsx,
    export_plan_de_baie_xlsx,
    export_synoptique_xlsx,
    export_ftth_json,
)
from .validate import export_validation
from .planner import export_plan_json


def _safe_name(s: str) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in (s or "ftth"))


def _export_all(proj, out_dir: str, tag: str, shape_dir: str | None = None) -> dict:
    """写出全部四类交付物(光交箱汇总/光路由表/机柜熔接盘图/系统图)+ 前端JSON + 自检报告，返回路径 dict。"""
    os.makedirs(out_dir, exist_ok=True)
    boite_path = os.path.join(out_dir, f"{tag}_Plans_de_Boite.xlsx")
    routes_path = os.path.join(out_dir, f"{tag}_Routes_Optiques.xlsx")
    export_plans_de_boite_xlsx(proj, boite_path)
    export_routes_optiques_xlsx(proj, routes_path)

    plan_de_baie = {}
    synoptique = {}
    for pm in proj.pm_codes():
        pm_safe = _safe_name(pm)
        pdb_path = os.path.join(out_dir, f"{tag}_Plan_de_Baie_{pm_safe}.xlsx")
        syn_path = os.path.join(out_dir, f"{tag}_Synoptique_{pm_safe}.xlsx")
        export_plan_de_baie_xlsx(proj, pdb_path, pm)
        export_synoptique_xlsx(proj, syn_path, pm)
        plan_de_baie[pm] = pdb_path
        synoptique[pm] = syn_path

    # 前端 S1 模块可用 JSON (箱体点位 + 光缆 + 汇总)
    ftth_json_path = os.path.join(out_dir, f"{tag}_ftth-data.json")
    export_ftth_json(proj, ftth_json_path)

    # 行业标准数据自检报告 (S3 校验规则复用)
    validation_path = os.path.join(out_dir, f"{tag}_ftth-validation.json")
    export_validation(proj, validation_path, shape_dir)

    # 正向智能规划设计产物 (反推重建 + 对比真实)
    plan_path = os.path.join(out_dir, f"{tag}_ftth-plan.json")
    export_plan_json(proj, plan_path, shape_dir=shape_dir)

    return {
        "project": proj,
        "boite_sommaire": boite_path,
        "routes_optiques": routes_path,
        "plan_de_baie": plan_de_baie,
        "synoptique": synoptique,
        "ftth_json": ftth_json_path,
        "validation": validation_path,
        "plan": plan_path,
        "summary": proj.summary(),
    }


def _export_all_single_workbook(proj, out_dir: str, tag: str, shape_dir: str | None = None) -> dict:
    """把 FTTH 四类交付物写入同一个 Excel 工作簿（多 sheet）。JSON / 自检报告仍单独输出。"""
    os.makedirs(out_dir, exist_ok=True)

    workbook_path = os.path.join(out_dir, f"{tag}_FTTH_Deliverables.xlsx")
    wb = openpyxl.Workbook()
    # 删除默认空白 sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 依次追加四类交付物（每张表一个或多个 sheet）
    export_routes_optiques_xlsx(proj, workbook=wb)
    export_plans_de_boite_xlsx(proj, workbook=wb)
    for pm in proj.pm_codes():
        export_plan_de_baie_xlsx(proj, pm_code=pm, workbook=wb)
        export_synoptique_xlsx(proj, pm_code=pm, workbook=wb)

    wb.save(workbook_path)

    # 前端 JSON 与自检报告仍保持独立文件
    ftth_json_path = os.path.join(out_dir, f"{tag}_ftth-data.json")
    export_ftth_json(proj, ftth_json_path)
    validation_path = os.path.join(out_dir, f"{tag}_ftth-validation.json")
    export_validation(proj, validation_path, shape_dir)
    plan_path = os.path.join(out_dir, f"{tag}_ftth-plan.json")
    export_plan_json(proj, plan_path, shape_dir=shape_dir)

    return {
        "project": proj,
        "workbook": workbook_path,
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "ftth_json": ftth_json_path,
        "validation": validation_path,
        "plan": plan_path,
        "summary": proj.summary(),
    }


def export_from_dbf(shape_dir: str, out_dir: str, prefix: str = "",
                    pm_filter: list[str] | None = None) -> dict:
    """从 Shape 目录装载并导出四类交付物，返回输出路径 dict。

    pm_filter: 可选 PM 编码列表，仅导出归属这些 PM 的局部成果(文件名追加 _PM 后缀)。
    """
    proj = load_dbf(shape_dir, pm_filter=pm_filter)
    tag = _safe_name(prefix or os.path.basename(os.path.normpath(shape_dir)))
    if pm_filter:
        tag += "_" + "_".join(_safe_name(p) for p in sorted(pm_filter))
    return _export_all(proj, out_dir, tag, shape_dir=shape_dir)


def export_from_dbf_single_workbook(shape_dir: str, out_dir: str, prefix: str = "",
                                    pm_filter: list[str] | None = None) -> dict:
    """从 Shape 目录装载并导出为单个 Excel 工作簿（多 sheet）。

    与 export_from_dbf 的区别：
    - 不生成多个独立 xlsx；
    - 光路由表 / 光交箱汇总 / 机柜熔接盘图 / 系统图 全部写入一个 xlsx 的不同 sheet。
    """
    proj = load_dbf(shape_dir, pm_filter=pm_filter)
    tag = _safe_name(prefix or os.path.basename(os.path.normpath(shape_dir)))
    if pm_filter:
        tag += "_" + "_".join(_safe_name(p) for p in sorted(pm_filter))
    return _export_all_single_workbook(proj, out_dir, tag, shape_dir=shape_dir)


def export_from_qgis(layers, out_dir: str, prefix: str = "qgis") -> dict:
    """从 QGIS 图层装载并导出四类交付物。"""
    proj = load_qgis(layers)
    tag = _safe_name(prefix)
    return _export_all(proj, out_dir, tag, shape_dir=None)
