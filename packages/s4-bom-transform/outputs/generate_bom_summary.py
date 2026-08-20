"""
生成项目级 BOM 汇总 Excel（模仿用户截图格式）
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 路径 ──
BASE = Path(__file__).resolve().parent.parent
CATALOG_PATH = BASE / "s4-bom-transform" / "engine" / "data" / "material_catalog.json"
DESIGN_PATH  = BASE / "s4-bom-transform" / "engine" / "data" / "mock" / "design_yuncheng_site_A001.json"
OUTPUT_PATH  = BASE / "outputs" / "物料清单_BOM_运城5G基站项目.xlsx"

# ── 参考单价（元）──
PRICE_MAP = {
    "M-ANT-001": 45000, "M-ANT-002": 800,   "M-ANT-003": 200,   "M-ANT-004": 300,
    "M-ANT-005": 12000,
    "M-RRU-001": 18000,
    "M-BBU-001": 25000, "M-BBU-002": 15000,
    "M-PWR-001": 12000, "M-PWR-002": 8000,  "M-PWR-003": 3500,  "M-PWR-004": 600,
    "M-TRN-001": 22000, "M-TRN-002": 2800,
    "M-RACK-001": 5500,
    "M-IRU-001": 8000,  "M-IHU-001": 6000,
    "M-MRU-001": 10000, "M-MHU-001": 5000,
    "M-ACC-001": 150,   "M-ACC-002": 80,    "M-ACC-003": 120,   "M-ACC-004": 5,
    "M-ACC-005": 120,   "M-ACC-006": 3,     "M-ACC-007": 25,    "M-ACC-008": 180,
    "M-ACC-009": 200,   "M-ACC-010": 80,    "M-ACC-011": 800,   "M-ACC-012": 300,
    "M-ACC-013": 15,    "M-ACC-014": 3,     "M-ACC-015": 300,   "M-ACC-016": 250,
    "M-ACC-017": 280,   "M-ACC-018": 120,   "M-ACC-019": 180,   "M-ACC-020": 3500,
    "M-ACC-021": 80,    "M-ACC-022": 200,   "M-ACC-023": 30,    "M-ACC-024": 50,
    "M-ACC-025": 60,    "M-ACC-026": 150,   "M-ACC-027": 120,
    "M-CBL-001": 80,    "M-CBL-002": 45,    "M-CBL-003": 70,    "M-CBL-004": 120,
    "M-CBL-005": 55,    "M-CBL-006": 35,    "M-CBL-007": 40,    "M-CBL-008": 25,
    "M-CBL-009": 60,    "M-CBL-010": 80,    "M-CBL-011": 50,    "M-CBL-012": 120,
    "M-CBL-013": 30,    "M-CBL-014": 150,
}

# 分类规则
def classify(code: str, name: str) -> str:
    if "机柜" in name or code.startswith("M-RACK") or code in ("M-ACC-015", "M-ACC-016", "M-ACC-017"):
        return "塔杆"
    if code.startswith("M-ANT") or code.startswith("M-IRU") or "天线" in name:
        return "天线"
    if "接地" in name or code in ("M-ACC-020",):
        return "接地"
    if code.startswith(("M-BBU", "M-RRU", "M-PWR", "M-TRN", "M-IHU", "M-MRU", "M-MHU")):
        return "设备"
    if code.startswith("M-ACC"):
        return "配套"
    if code.startswith("M-CBL"):
        return "馈线"
    return "配套"


def load_catalog():
    with open(CATALOG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def load_design():
    with open(DESIGN_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def build_bom_items(catalog, design):
    """根据设计清单和物料编码库生成完整 BOM 明细"""
    # 构建编码库查找表
    lookup = {}
    for m in catalog["mappings"]:
        key = (m["deviceType"], m["deviceModel"])
        lookup[key] = m

    items = []
    device_count = {}

    # 统计设备数量
    for dev in design["devices"]:
        key = (dev["type"], dev["model"])
        device_count[key] = device_count.get(key, 0) + dev.get("qty", 1)

    # 逐设备展开物料
    for (dtype, dmodel), qty in device_count.items():
        mapping = lookup.get((dtype, dmodel))
        if not mapping:
            continue

        # 主设备
        md = mapping["mainDevice"]
        items.append({
            "code": md["materialCode"], "name": md["materialName"],
            "spec": md.get("spec", ""), "unit": md["unit"],
            "qty": qty, "price": PRICE_MAP.get(md["materialCode"], 0)
        })

        # 辅材
        for aux in mapping.get("auxiliaries", []):
            items.append({
                "code": aux["materialCode"], "name": aux["materialName"],
                "spec": aux.get("spec", ""), "unit": aux["unit"],
                "qty": aux["qtyPerDevice"] * qty,
                "price": PRICE_MAP.get(aux["materialCode"], 0)
            })

        # 线缆（简化：使用固定长度估算）
        for cab in mapping.get("cables", []):
            # 简化长度估算
            length = 25.0  # 默认 25m
            if cab.get("calcMethod") == "fixed_3m":
                length = 3.0
            elif cab.get("calcMethod") == "fixed_15m":
                length = 15.0
            elif cab.get("calcMethod") == "tower_to_ground":
                length = 25.0
            elif cab.get("calcMethod") == "rack_ground":
                length = 3.0
            elif cab.get("calcMethod") == "short_ground":
                length = 2.0
            elif cab.get("calcMethod") == "rack_to_power":
                length = 5.0
            elif cab.get("calcMethod") == "battery_to_power":
                length = 3.0
            elif cab.get("calcMethod") == "dist_to_power_panel":
                length = 10.0
            elif cab.get("calcMethod") == "tower_top_to_bbu":
                length = 30.0

            cable_qty = cab["qtyPerDevice"] * qty
            items.append({
                "code": cab["materialCode"], "name": cab["materialName"],
                "spec": cab.get("spec", ""), "unit": cab["unit"],
                "qty": cable_qty, "price": PRICE_MAP.get(cab["materialCode"], 0),
                "length": round(length, 2)
            })

    # 站点级辅材
    for sa in catalog.get("siteLevelAuxiliaries", {}).get("items", []):
        qty = sa["qtyPerSite"]
        items.append({
            "code": sa["materialCode"], "name": sa["materialName"],
            "spec": sa.get("spec", ""), "unit": sa["unit"],
            "qty": qty, "price": PRICE_MAP.get(sa["materialCode"], 0)
        })

    return items


def aggregate_by_category(items):
    """按分类汇总：{分类: {count, amount}}"""
    agg = {}
    for it in items:
        cat = classify(it["code"], it["name"])
        price = it["price"]
        qty = it["qty"]
        total = price * qty
        if cat not in agg:
            agg[cat] = {"count": 0, "amount": 0}
        agg[cat]["count"] += qty
        agg[cat]["amount"] += total
    return agg


def main():
    catalog = load_catalog()
    design = load_design()

    items = build_bom_items(catalog, design)
    agg = aggregate_by_category(items)

    total_items = sum(v["count"] for v in agg.values())
    total_amount = sum(v["amount"] for v in agg.values())

    # ── Excel 生成 ──
    wb = Workbook()
    ws = wb.active
    ws.title = "物料清单汇总"

    # 样式定义
    title_font = Font(name="微软雅黑", size=16, bold=True, color="1B5E20")
    title_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    title_border = Border(
        left=Side(style="thin", color="2E7D32"),
        right=Side(style="thin", color="2E7D32"),
        top=Side(style="thin", color="2E7D32"),
        bottom=Side(style="thin", color="2E7D32"),
    )

    info_label_font = Font(name="微软雅黑", size=11, bold=True)
    info_value_font = Font(name="微软雅黑", size=11)
    info_align = Alignment(vertical="center")

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    data_font = Font(name="微软雅黑", size=11)
    data_align = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="388E3C"),
        right=Side(style="thin", color="388E3C"),
        top=Side(style="thin", color="388E3C"),
        bottom=Side(style="thin", color="388E3C"),
    )

    # 1. 标题行（A1:D1 合并）
    ws.merge_cells("A1:D1")
    ws["A1"] = "物料清单（BOM）— 运城5G基站建设项目"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = title_border
    ws.row_dimensions[1].height = 36

    # 2. 项目信息区
    info_data = [
        ("项目编号", "PRJ-YUNCHENG-5G-01"),
        ("项目名称", "运城5G基站建设项目"),
        ("创建时间", "2026-08-09 14:30:00"),
        ("创建人", "PersonB"),
        ("站点数量", 1),
        ("物料条目数", len(items)),
        ("物料总金额", total_amount),
        ("状态", "draft"),
    ]

    for i, (label, value) in enumerate(info_data, start=3):
        ws.cell(row=i, column=1, value=label).font = info_label_font
        ws.cell(row=i, column=1).alignment = info_align
        if label == "物料总金额":
            ws.cell(row=i, column=2, value=f"¥ {value:,.2f}").font = info_value_font
        else:
            ws.cell(row=i, column=2, value=value).font = info_value_font
        ws.cell(row=i, column=2).alignment = info_align

    # 3. 分类汇总区
    summary_start = 12
    ws.merge_cells(start_row=summary_start, start_column=1, end_row=summary_start, end_column=3)
    ws.cell(row=summary_start, column=1, value="分类汇总").font = Font(name="微软雅黑", size=13, bold=True)

    # 表头
    headers = ["分类", "条目数", "金额(¥)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=summary_start + 1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 分类顺序：塔杆、天线、接地、设备、配套、馈线
    cat_order = ["塔杆", "天线", "接地", "设备", "配套", "馈线"]
    row_idx = summary_start + 2
    for cat in cat_order:
        if cat not in agg:
            continue
        data = agg[cat]
        vals = [cat, data["count"], data["amount"]]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if col == 3:
                cell.number_format = '#,##0'
        row_idx += 1

    # 列宽
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    print(f"Excel 已生成: {OUTPUT_PATH}")
    print(f"物料条目总数: {len(items)}")
    print(f"物料总金额: ¥{total_amount:,.2f}")
    print("\n分类汇总:")
    for cat in cat_order:
        if cat in agg:
            print(f"  {cat}: {agg[cat]['count']} 条, ¥{agg[cat]['amount']:,.2f}")


if __name__ == "__main__":
    main()
