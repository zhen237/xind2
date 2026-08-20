"""
生成三个场景的项目级 BOM 汇总 Excel
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path(__file__).resolve().parent.parent
CATALOG_PATH = BASE / "s4-bom-transform" / "engine" / "data" / "material_catalog.json"
MOCK_DIR     = BASE / "s4-bom-transform" / "engine" / "data" / "mock"
OUTPUT_DIR   = BASE / "outputs"

SCENARIOS = [
    {
        "key": "D001",
        "mock_file": "design_yuncheng_site_A001.json",
        "output": "物料清单_BOM_运城5G宏站.xlsx",
        "project_name": "运城5G宏站建设项目",
        "project_id": "PRJ-YUNCHENG-5G-01",
        "creator": "PersonB",
    },
    {
        "key": "D002",
        "mock_file": "design_indoor_B001.json",
        "output": "物料清单_BOM_万象城室分.xlsx",
        "project_name": "万象城商业综合体室分覆盖",
        "project_id": "PRJ-indoor-mall",
        "creator": "S1设计引擎",
    },
    {
        "key": "D003",
        "mock_file": "design_micro_C001.json",
        "output": "物料清单_BOM_解放路微站.xlsx",
        "project_name": "市区补盲微站覆盖项目",
        "project_id": "PRJ-micro-urban",
        "creator": "S1设计引擎",
    },
]

PRICE_MAP = {
    "M-ANT-001": 45000, "M-ANT-002": 800,   "M-ANT-003": 200,   "M-ANT-004": 300,
    "M-ANT-005": 12000,
    "M-RRU-001": 18000,
    "M-BBU-001": 25000, "M-BBU-002": 15000, "M-BBU-003": 22000,
    "M-PWR-001": 12000, "M-PWR-002": 8000,  "M-PWR-003": 3500,  "M-PWR-004": 600,
    "M-PWR-005": 10000,
    "M-TRN-001": 22000, "M-TRN-002": 2800,  "M-TRN-003": 20000,
    "M-RACK-001": 5500,  "M-RACK-002": 6500, "M-RACK-003": 4800,
    "M-IRU-001": 8000,  "M-IHU-001": 6000,
    "M-MRU-001": 10000, "M-MHU-001": 5000,
    "M-ACC-001": 150,   "M-ACC-002": 80,    "M-ACC-003": 120,   "M-ACC-004": 5,
    "M-ACC-005": 120,   "M-ACC-006": 3,     "M-ACC-007": 25,    "M-ACC-008": 180,
    "M-ACC-009": 200,   "M-ACC-010": 80,    "M-ACC-011": 800,   "M-ACC-012": 300,
    "M-ACC-013": 15,    "M-ACC-014": 3,     "M-ACC-015": 300,   "M-ACC-016": 250,
    "M-ACC-017": 280,   "M-ACC-018": 120,   "M-ACC-019": 180,   "M-ACC-020": 3500,
    "M-ACC-021": 80,    "M-ACC-022": 200,   "M-ACC-023": 30,    "M-ACC-024": 50,
    "M-ACC-025": 60,    "M-ACC-026": 150,   "M-ACC-027": 120,
    "M-CBL-001": 80,    "M-CBL-002": 45,    "M-CBL-003": 70,    "M-CBL-004": 120,
    "M-CBL-005": 55,    "M-CBL-006": 35,    "M-CBL-007": 40,    "M-CBL-008": 25,
    "M-CBL-009": 60,    "M-CBL-010": 80,    "M-CBL-011": 50,    "M-CBL-012": 120,
    "M-CBL-013": 30,    "M-CBL-014": 150,
}


def classify(code: str, name: str) -> str:
    if "机柜" in name or code.startswith("M-RACK") or code in ("M-ACC-015", "M-ACC-016", "M-ACC-017"):
        return "塔杆"
    if code.startswith("M-ANT") or code.startswith("M-IRU") or "天线" in name:
        return "天线"
    if "接地" in name or code in ("M-ACC-020",):
        return "接地"
    if code.startswith(("M-BBU", "M-RRU", "M-PWR", "M-TRN", "M-IHU", "M-MRU", "M-MHU")):
        return "设备"
    if code.startswith("M-ACC"):
        return "配套"
    if code.startswith("M-CBL"):
        return "馈线"
    return "配套"


def normalize_device(dev: dict) -> dict:
    """兼容两种 mock 数据格式"""
    d = {}
    d["type"] = dev.get("type") or dev.get("deviceType", "")
    d["model"] = dev.get("model") or dev.get("deviceModel", "")
    d["qty"] = dev.get("qty", 1)
    coords = dev.get("coordinates", {})
    d["lat"] = coords.get("lat") or dev.get("latitude", 0)
    d["lng"] = coords.get("lng") or dev.get("longitude", 0)
    d["alt"] = coords.get("alt") or dev.get("altitude", 0)
    return d


CABLE_LENGTH = {
    "fixed_3m": 3.0, "fixed_15m": 15.0, "tower_to_ground": 25.0,
    "rack_ground": 3.0, "short_ground": 2.0, "rack_to_power": 5.0,
    "battery_to_power": 3.0, "dist_to_power_panel": 10.0,
    "tower_top_to_bbu": 30.0, "horizontal_distance_x1.2": 25.0,
    "distance_plus_riser": 30.0,
}


def build_items(catalog, design):
    lookup = {}
    for m in catalog["mappings"]:
        key = (m["deviceType"], m["deviceModel"])
        lookup[key] = m

    devices = design.get("devices", [])
    device_count = {}
    for raw in devices:
        dev = normalize_device(raw)
        key = (dev["type"], dev["model"])
        device_count[key] = device_count.get(key, 0) + dev["qty"]

    items = []
    for (dtype, dmodel), qty in device_count.items():
        mapping = lookup.get((dtype, dmodel))
        if not mapping:
            continue

        md = mapping["mainDevice"]
        items.append({
            "code": md["materialCode"], "name": md["materialName"],
            "spec": md.get("spec", ""), "unit": md["unit"],
            "qty": qty, "price": PRICE_MAP.get(md["materialCode"], 0)
        })

        for aux in mapping.get("auxiliaries", []):
            items.append({
                "code": aux["materialCode"], "name": aux["materialName"],
                "spec": aux.get("spec", ""), "unit": aux["unit"],
                "qty": aux["qtyPerDevice"] * qty,
                "price": PRICE_MAP.get(aux["materialCode"], 0)
            })

        for cab in mapping.get("cables", []):
            length = CABLE_LENGTH.get(cab.get("calcMethod", ""), 25.0)
            items.append({
                "code": cab["materialCode"], "name": cab["materialName"],
                "spec": cab.get("spec", ""), "unit": cab["unit"],
                "qty": cab["qtyPerDevice"] * qty,
                "price": PRICE_MAP.get(cab["materialCode"], 0),
                "length": length
            })

    for sa in catalog.get("siteLevelAuxiliaries", {}).get("items", []):
        items.append({
            "code": sa["materialCode"], "name": sa["materialName"],
            "spec": sa.get("spec", ""), "unit": sa["unit"],
            "qty": sa["qtyPerSite"], "price": PRICE_MAP.get(sa["materialCode"], 0)
        })

    return items


def generate_excel(scenario, catalog, design):
    items = build_items(catalog, design)
    agg = {}
    for it in items:
        cat = classify(it["code"], it["name"])
        amt = it["price"] * it["qty"]
        if cat not in agg:
            agg[cat] = {"count": 0, "amount": 0}
        agg[cat]["count"] += it["qty"]
        agg[cat]["amount"] += amt

    total_count = sum(v["count"] for v in agg.values())
    total_amount = sum(v["amount"] for v in agg.values())

    wb = Workbook()
    ws = wb.active
    ws.title = "物料清单汇总"

    title_font = Font(name="微软雅黑", size=16, bold=True, color="1B5E20")
    title_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    title_border = Border(
        left=Side(style="thin", color="2E7D32"), right=Side(style="thin", color="2E7D32"),
        top=Side(style="thin", color="2E7D32"), bottom=Side(style="thin", color="2E7D32"),
    )
    label_font = Font(name="微软雅黑", size=11, bold=True)
    val_font = Font(name="微软雅黑", size=11)
    info_align = Alignment(vertical="center")
    hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="微软雅黑", size=11)
    data_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="388E3C"), right=Side(style="thin", color="388E3C"),
        top=Side(style="thin", color="388E3C"), bottom=Side(style="thin", color="388E3C"),
    )

    ws.merge_cells("A1:D1")
    title = f"物料清单（BOM）— {scenario['project_name']}"
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = title_border
    ws.row_dimensions[1].height = 36

    info = [
        ("项目编号", scenario["project_id"]),
        ("项目名称", scenario["project_name"]),
        ("创建时间", "2026-08-11 12:00:00"),
        ("创建人", scenario["creator"]),
        ("站点数量", design.get("siteMeta", {}).get("siteCount", 1)),
        ("物料条目数", len(items)),
        ("物料总金额", f"¥ {total_amount:,.2f}"),
        ("状态", "draft"),
    ]
    for i, (label, value) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=1).alignment = info_align
        ws.cell(row=i, column=2, value=value).font = val_font
        ws.cell(row=i, column=2).alignment = info_align

    sr = 12
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=3)
    ws.cell(row=sr, column=1, value="分类汇总").font = Font(name="微软雅黑", size=13, bold=True)

    headers = ["分类", "条目数", "金额(¥)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=sr + 1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = thin_border

    cat_order = ["塔杆", "天线", "接地", "设备", "配套", "馈线"]
    row_idx = sr + 2
    for cat in cat_order:
        if cat not in agg:
            continue
        vals = [cat, agg[cat]["count"], agg[cat]["amount"]]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.font = data_font; cell.alignment = data_align; cell.border = thin_border
            if col == 3:
                cell.number_format = '#,##0'
        row_idx += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10

    out_path = OUTPUT_DIR / scenario["output"]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path, len(items), total_amount, agg


def main():
    with open(CATALOG_PATH, "r", encoding="utf-8") as f:
        catalog = json.load(f)

    results = []
    for sc in SCENARIOS:
        mock_path = MOCK_DIR / sc["mock_file"]
        with open(mock_path, "r", encoding="utf-8") as f:
            design = json.load(f)
        path, count, amt, agg = generate_excel(sc, catalog, design)
        results.append((sc, path, count, amt, agg))

    for sc, path, count, amt, agg in results:
        print(f"\n===== {sc['project_name']} =====")
        print(f"  文件: {path}")
        print(f"  物料条目: {count}, 总金额: ¥{amt:,.2f}")
        for cat in ["塔杆", "天线", "接地", "设备", "配套", "馈线"]:
            if cat in agg:
                print(f"    {cat}: {agg[cat]['count']}条, ¥{agg[cat]['amount']:,.2f}")
    print("\n=== 三场景全部生成完成 ===")


if __name__ == "__main__":
    main()
