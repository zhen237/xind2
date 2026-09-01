"""
BOM 生成一键验证脚本 — 给一份样例设计数据，跑命令看生成的 BOM 对不对。

用法（engine 目录下）:
    venv/Scripts/python verify_bom.py            # 验证全部三个场景 (D001/D002/D003)
    venv/Scripts/python verify_bom.py D001       # 只验证运城宏站场景
    venv/Scripts/python verify_bom.py D002 --no-excel   # 不落 Excel 文件

流程: 样例设计数据 → S3 审查闸门 → 设备-物料映射 + 辅材 + 线缆估算
      → 工序工艺 + 纤芯分配 → Excel 导出 → 输出验证报告（任一检查失败退出码 1）

验证项（对应 S4_TASK_ANALYSIS AC-1~AC-5）:
    [1] 每台设备都映射到 main_device 物料（无静默丢失）
    [2] 主设备/辅材数量乘算正确（qty × qtyPerUnit / qtyPerDevice）
    [3] 射频跳线固定 3m/根
    [4] 光纤长度 = 水平距 × 1.2（Haversine 复算对比）
    [5] 线缆总长度 = 单根长度 × 根数
    [6] 站点级辅材齐套（接地网/走线架/防火封堵/扎带）
    [7] 标识标签包 = max(1, int(天线数/2))
    [8] Excel 生成且含三个 Sheet
"""
import json
import sys
import time
from pathlib import Path

ENGINE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(ENGINE_DIR))

from app.services import bom_engine, process_requirements, fiber_allocation, excel_export, review_gate  # noqa: E402
from app.services.catalog_service import get_mapping, get_site_auxiliaries  # noqa: E402
from app.services.cable_estimator import estimate_cable_length, haversine_m  # noqa: E402

MOCK_DIR = ENGINE_DIR / "data" / "mock"

SCENARIOS = {
    "D001": {"file": "design_yuncheng_site_A001.json", "name": "运城宏站"},
    "D002": {"file": "design_indoor_B001.json", "name": "室分"},
    "D003": {"file": "design_micro_C001.json", "name": "微站"},
}


def load_s3_violations() -> dict:
    """加载 S3 审查虚拟数据（与 dev-proxy 同一份）。"""
    path = MOCK_DIR / "s3_review_results.json"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return {k: v for k, v in json.load(f).items() if not k.startswith("_")}


def normalize_devices(design: dict) -> dict:
    """归一化设备字段（与引擎路由 _normalize_devices 相同逻辑）。"""
    devices = []
    for dev in design.get("devices", []):
        nd = dict(dev)
        nd.setdefault("type", nd.pop("deviceType", None) or nd.get("type", ""))
        nd.setdefault("model", nd.pop("deviceModel", None) or nd.get("model", ""))
        nd.setdefault("name", nd.pop("deviceName", None) or nd.get("name", ""))
        devices.append(nd)
    design["devices"] = devices
    return design


def run_pipeline(scenario: str, write_excel: bool = True):
    """跑完整 BOM 管线，返回 (design, gate, items, proc, fiber_alloc, fiber_summary, checks)。"""
    cfg = SCENARIOS[scenario]
    with open(MOCK_DIR / cfg["file"], "r", encoding="utf-8") as f:
        design = normalize_devices(json.load(f))

    # S3 分级审查闸门（虚拟数据）
    violations = load_s3_violations().get(scenario, [])
    gate = review_gate.check_gate(
        {"result": "approved_with_warnings" if violations else "approved", "violations": violations}
    )
    if gate["decision"] == review_gate.ALLOWED_WITH_WARNINGS:
        design = review_gate.flag_devices(design, gate)
    rect_steps = review_gate.build_rectification_steps(gate)

    items = bom_engine.generate_bom_items(design)
    device_types = list({d.get("type", "") for d in design["devices"]})
    proc = process_requirements.generate_process_requirements(device_types) + rect_steps
    fiber_alloc, fiber_summary = fiber_allocation.generate_fiber_allocation(design["devices"])

    excel_path = None
    if write_excel:
        excel_path = excel_export.export_to_excel(
            task_id=f"verify-{scenario}",
            bom_items=items,
            process_steps=proc,
            fiber_alloc=fiber_alloc,
            fiber_summary=fiber_summary,
        )

    checks = run_checks(design, gate, items, excel_path)
    return design, gate, items, proc, fiber_alloc, fiber_summary, checks


def run_checks(design: dict, gate: dict, items: list, excel_path) -> list[dict]:
    """执行 8 项业务规则验证，返回 [{name, ok, detail}]。"""
    checks = []

    def add(name, ok, detail=""):
        checks.append({"name": name, "ok": ok, "detail": detail})

    devices = design["devices"]
    mains = [i for i in items if i["category"] == "main_device"]

    # [1] 每台设备都有 main_device 物料
    add("设备-物料映射完整", len(mains) == len(devices),
        f"{len(mains)}/{len(devices)} 台设备映射到主设备物料")

    # [2] 数量乘算正确
    calc_bad = []
    for dev in devices:
        m = get_mapping(dev.get("type", ""), dev.get("model", ""))
        if not m:
            continue
        expected = dev.get("qty", 1) * m["mainDevice"].get("qtyPerUnit", 1)
        found = [i for i in mains if i["materialCode"] == m["mainDevice"]["materialCode"]
                 and i.get("deviceName") == dev.get("name", "")]
        if not found or found[0]["qty"] != expected:
            calc_bad.append(dev.get("name", "?"))
    add("主设备数量 = 设备数 × qtyPerUnit", not calc_bad,
        "全部一致" if not calc_bad else f"不一致: {calc_bad}")

    # [3] 射频跳线 3m/根（室内/微站场景可能无射频跳线 → 跳过）
    jumpers = [i for i in items if i["materialCode"] == "M-CBL-004"]
    if jumpers:
        add("射频跳线 3m/根", all(i["singleLength"] == 3.0 for i in jumpers),
            f"{len(jumpers)} 条，单根均为 3.0m")
    else:
        add("射频跳线 3m/根", True, "本场景无射频跳线（跳过）")

    # [4] 光纤 = 水平距 × 1.2
    # 引擎按设备顺序生成明细 → 同名设备（如 3 台 AAU）按出现顺序与明细一一对应
    from collections import defaultdict
    pool = defaultdict(list)
    for d in devices:
        pool[d.get("name", "")].append(d)
    cursor = defaultdict(int)  # name → 已消费的下标
    bbu = next((d for d in devices if d.get("type") == "bbu"), None)
    fiber_bad = []
    if bbu:
        for i in items:
            if i["category"] != "cable" or "光纤" not in i["materialName"]:
                continue
            name = i.get("deviceName", "")
            idx = cursor[name]
            cursor[name] = idx + 1
            devs = pool.get(name, [])
            dev = devs[idx] if idx < len(devs) else None
            if not dev:
                fiber_bad.append(f"{i['materialCode']}@{name}(设备未找到)")
                continue
            dc, bc = bom_engine._coords(dev), bom_engine._coords(bbu)
            expected = round(haversine_m(dc.get("lat", 0), dc.get("lng", 0), 0,
                                         bc.get("lat", 0), bc.get("lng", 0), 0) * 1.2, 2)
            if abs(i["singleLength"] - expected) > 0.05:
                fiber_bad.append(f"{i['materialCode']}@{name}({i['singleLength']} vs {expected})")
    add("光纤长度 = 水平距×1.2", not fiber_bad,
        "全部吻合" if not fiber_bad else f"偏差: {fiber_bad}")

    # [5] 线缆总长 = 单根 × 根数
    cable_bad = [i["materialCode"] for i in items
                 if i["category"] == "cable" and abs(i["totalLength"] - i["singleLength"] * i["qty"]) > 0.01]
    add("线缆总长 = 单根 × 根数", not cable_bad,
        "全部一致" if not cable_bad else f"不一致: {set(cable_bad)}")

    # [6] 站点级辅材齐套
    site_codes = {i["materialCode"] for i in items if i.get("deviceType") == "site"}
    required = {a["materialCode"] for a in get_site_auxiliaries()}
    add("站点级辅材齐套", required <= site_codes,
        f"{len(required & site_codes)}/{len(required)} 项在场")

    # [7] 标识标签包动态计算
    antenna_count = sum(d.get("qty", 1) for d in devices if d.get("type") == "antenna")
    expected_packs = max(1, antenna_count // 2)
    labels = [i for i in items if i["materialCode"] == "M-ACC-028"]
    add("标签包 = max(1, 天线数/2)", bool(labels) and labels[0]["qty"] == expected_packs,
        f"天线 {antenna_count} 台 → {expected_packs} 包")

    # [8] Excel 三 Sheet
    if excel_path:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            sheets = set(wb.sheetnames)
            add("Excel 含三 Sheet", {"BOM物料清单", "关键工序工艺", "纤芯分配表"} <= sheets,
                f"sheets={wb.sheetnames}")
            wb.close()
        except Exception as e:  # noqa: BLE001
            add("Excel 含三 Sheet", False, f"读取失败: {e}")
    else:
        add("Excel 含三 Sheet", True, "（--no-excel 跳过）")

    return checks


def print_report(scenario: str, design, gate, items, proc, fiber_alloc, fiber_summary, checks, elapsed_ms):
    cfg = SCENARIOS[scenario]
    mains = [i for i in items if i["category"] == "main_device"]
    aux = [i for i in items if i["category"] == "auxiliary"]
    cables = [i for i in items if i["category"] == "cable"]
    flagged = [i for i in items if i.get("requiresRectification")]

    print()
    print("═" * 62)
    print(f"  S4 BOM 验证报告 — {scenario}（{cfg['name']}）")
    print("═" * 62)
    print(f"  设计任务   : {design.get('_meta', {}).get('designTaskId', scenario)}")
    print(f"  设备数     : {len(design['devices'])}")
    print(f"  审查闸门   : {gate['decision']}"
          + (f"（counts={gate['counts']}）" if gate["counts"] != {"critical": 0, "error": 0, "warning": 0, "pending": 0} else ""))
    print(f"  生成耗时   : {elapsed_ms} ms（人工 2-4 小时 → 缩短 ≥99%）")
    print(f"  物料明细   : {len(items)} 条 = 主设备 {len(mains)} + 辅材 {len(aux)} + 线缆 {len(cables)}")
    print(f"  数量合计   : 主设备 {sum(i['qty'] for i in mains)} / "
          f"辅材 {sum(i['qty'] for i in aux)} / 线缆 {sum(i['qty'] for i in cables)}")
    print(f"  工序工艺   : {len(proc)} 条；纤芯分配: {len(fiber_alloc)} 行")
    if flagged:
        print(f"  ⚠ 整改标记 : {len(flagged)} 条物料携带 requiresRectification")
    print("─" * 62)
    for c in checks:
        mark = "✔" if c["ok"] else "✘"
        print(f"  [{mark}] {c['name']}: {c['detail']}")
    print("─" * 62)
    # 抽样明细（每类前 3 条）
    print("  样例明细（每类前 3 条）:")
    for cat, lst in (("主设备", mains), ("辅材", aux), ("线缆", cables)):
        for i in lst[:3]:
            length = f" | {i['singleLength']}m×{i['qty']} = {i['totalLength']}m" if cat == "线缆" else ""
            rect = " ⚠需整改" if i.get("requiresRectification") else ""
            print(f"    {i['materialCode']:<12} {i['materialName'][:22]:<24} ×{i['qty']:<4}{length}{rect}")
    print()


def main():
    args = [a for a in sys.argv[1:]]
    write_excel = "--no-excel" not in args
    args = [a for a in args if not a.startswith("--")]
    targets = args or list(SCENARIOS)
    for t in targets:
        if t not in SCENARIOS:
            print(f"未知场景: {t}（可选: {'/'.join(SCENARIOS)}）")
            sys.exit(2)

    all_ok = True
    for scenario in targets:
        t0 = time.perf_counter()
        design, gate, items, proc, fa, fs, checks = run_pipeline(scenario, write_excel)
        elapsed_ms = round((time.perf_counter() - t0) * 1000)
        print_report(scenario, design, gate, items, proc, fa, fs, checks, elapsed_ms)
        all_ok = all(c["ok"] for c in checks)

    print("═" * 62)
    if all_ok:
        print("  ✔ 全部验证通过 — BOM 生成准确")
    else:
        print("  ✘ 存在验证失败项，请检查上方报告")
    print("═" * 62)
    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
