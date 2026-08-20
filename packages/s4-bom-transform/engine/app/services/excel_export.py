"""
Excel 导出服务 — openpyxl 生成 .xlsx，BOM + 工序工艺 + 纤芯分配三 sheet。
"""
import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger("s4-engine.export")

EXPORT_DIR = Path(__file__).resolve().parent.parent.parent / "exports"


def export_to_excel(task_id: str,
                    bom_items: list[dict],
                    process_steps: Optional[list[dict]] = None,
                    fiber_alloc: Optional[list[dict]] = None,
                    fiber_summary: Optional[dict] = None) -> str:
    """
    生成综合施工指令 Excel 文件（.xlsx）。

    3 个 Sheet：
      - BOM 物料清单（主设备/辅材/线缆）
      - 关键工序工艺
      - 纤芯分配表

    返回: 文件绝对路径
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    filepath = EXPORT_DIR / f"{task_id}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    # ───── 公共样式 ─────
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, headers: list[str]):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_row(ws, row_idx: int, values: list):
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    # ───── Sheet 1: BOM 物料清单 ─────
    ws1 = wb.create_sheet("BOM物料清单")
    bom_headers = ["序号", "物料编码", "物料名称", "规格型号", "单位", "数量",
                   "类别", "单根长度(m)", "总长度(m)", "关联设备"]
    style_header(ws1, bom_headers)

    # 颜色标记不同类别
    cat_fills = {
        "main_device": PatternFill(start_color="E3F0FF", end_color="E3F0FF", fill_type="solid"),
        "auxiliary": PatternFill(start_color="FFF9E3", end_color="FFF9E3", fill_type="solid"),
        "cable": PatternFill(start_color="E3FFE3", end_color="E3FFE3", fill_type="solid"),
    }

    for i, item in enumerate(bom_items, 1):
        cat = item.get("category", "")
        values = [
            i,
            item.get("materialCode", ""),
            item.get("materialName", ""),
            item.get("spec", ""),
            item.get("unit", ""),
            item.get("qty", 0),
            {"main_device": "主设备", "auxiliary": "辅材", "cable": "线缆"}.get(cat, cat),
            item.get("singleLength", ""),
            item.get("totalLength", ""),
            item.get("deviceName", ""),
        ]
        style_row(ws1, i + 1, values)
        fill = cat_fills.get(cat)
        if fill:
            for col in range(1, len(bom_headers) + 1):
                ws1.cell(row=i + 1, column=col).fill = fill

    # 列宽
    for col_idx, w in enumerate([6, 14, 32, 32, 6, 8, 10, 14, 14, 24], 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = w

    # ───── Sheet 2: 关键工序工艺 ─────
    if process_steps:
        ws2 = wb.create_sheet("关键工序工艺")
        proc_headers = ["序号", "工序名称", "工艺要求", "验收标准", "适用设备类型"]
        style_header(ws2, proc_headers)
        for i, step in enumerate(process_steps, 1):
            values = [
                step.get("序号", i),
                step.get("工序名称", ""),
                step.get("工艺要求", ""),
                step.get("验收标准", ""),
                step.get("适用设备类型", ""),
            ]
            style_row(ws2, i + 1, values)
        for col_idx, w in enumerate([6, 20, 55, 55, 20], 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = w

    # ───── Sheet 3: 纤芯分配表 ─────
    if fiber_alloc:
        ws3 = wb.create_sheet("纤芯分配表")
        fiber_headers = ["ODF端口", "纤芯号", "起始设备", "起始端口", "终止设备",
                         "终止端口", "纤芯类型", "纤芯用途", "长度(m)"]
        style_header(ws3, fiber_headers)
        for i, fa in enumerate(fiber_alloc, 1):
            values = [
                fa.get("ODF端口", ""),
                fa.get("纤芯号", ""),
                fa.get("起始设备", ""),
                fa.get("起始端口", ""),
                fa.get("终止设备", ""),
                fa.get("终止端口", ""),
                fa.get("纤芯类型", ""),
                fa.get("纤芯用途", ""),
                fa.get("长度(m)", ""),
            ]
            style_row(ws3, i + 1, values)

        # 汇总行
        if fiber_summary:
            summary_row = len(fiber_alloc) + 2
            ws3.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
            ws3.cell(row=summary_row, column=1, value=f"纤芯使用: {fiber_summary['total_cores_assigned']}/"
                      f"{fiber_summary['odf_capacity']} ({fiber_summary['odf_usage_rate']})"
                      f" | 预留: {fiber_summary['reserve_cores']}").font = Font(bold=True)
            ws3.merge_cells(start_row=summary_row+1, start_column=1, end_row=summary_row+1, end_column=9)
            ws3.cell(row=summary_row+1, column=1,
                     value="注: 长度由线缆估算模块自动填入，导出时为占位符").font = Font(italic=True, color="888888")

        for col_idx, w in enumerate([10, 8, 28, 14, 28, 14, 18, 32, 18], 1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(str(filepath))
    logger.info(f"Excel exported: {filepath}")

    return str(filepath)
